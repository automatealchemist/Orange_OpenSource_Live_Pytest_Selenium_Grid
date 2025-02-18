import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HELLO\\Downloads\\login_data.xlsx")

sheet = book.active
cell = sheet.cell(row=1,column=2)
#print(cell.value)
dict ={}
for i in range(1,sheet.max_row+1):
    for j in range(1, sheet.max_column + 1):
        #print(sheet.cell(row=i,column=j).value)
        #dict["username"]={}
        dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(dict)